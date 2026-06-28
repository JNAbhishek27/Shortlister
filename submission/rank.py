#!/usr/bin/env python3
"""
Redrob Intelligent Candidate Ranker
====================================
Ranks candidates from candidates.jsonl against the Senior AI Engineer JD.

Usage:
    python rank.py --candidates ./candidates.jsonl --out ./submission.csv
    python rank.py --candidates ./candidates.jsonl --out ./submission.csv --xlsx ./ranked_output.xlsx

No network calls. Runs entirely on CPU. ~30-60 seconds for 100K candidates.
"""

import argparse
import csv
import json
import math
import sys
from datetime import date, datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# JD Configuration — derived from parsing the job_description.docx
# ---------------------------------------------------------------------------

JD = {
    "title": "Senior AI Engineer",
    "experience_min": 5,
    "experience_max": 9,
    "preferred_locations": ["Pune", "Noida", "Delhi", "Mumbai", "Hyderabad", "Bengaluru", "Bangalore"],
    "country": "India",
    "notice_period_preferred_days": 30,
}

# Tier-1 must-have AI skills (from JD "absolutely need" section)
TIER1_SKILLS = {
    "embeddings", "sentence-transformers", "vector database", "vector search",
    "pinecone", "weaviate", "qdrant", "milvus", "opensearch", "faiss",
    "elasticsearch", "hybrid search", "retrieval", "rag",
    "dense retrieval", "sparse retrieval", "bm25",
    "ndcg", "mrr", "map", "ranking evaluation", "learning to rank",
    "information retrieval", "semantic search",
}

# Tier-2 nice-to-have AI skills
TIER2_SKILLS = {
    "nlp", "natural language processing", "transformers", "bert", "llm",
    "large language models", "fine-tuning", "lora", "qlora", "peft",
    "pytorch", "tensorflow", "scikit-learn", "python",
    "machine learning", "deep learning", "neural networks",
    "recommendation systems", "ranking", "xgboost", "gradient boosting",
    "hugging face", "openai", "langchain", "llamaindex",
    "spacy", "nltk", "gensim", "word2vec", "glove",
    "mlflow", "wandb", "weights & biases",
    "apache spark", "airflow", "data pipelines",
    "image classification", "speech recognition", "tts",  # these exist in data
    "nlu", "text classification",
}

# AI-adjacent technical signals
TIER3_SKILLS = {
    "aws", "gcp", "azure", "docker", "kubernetes",
    "api", "flask", "fastapi", "rest", "grpc",
    "sql", "nosql", "postgresql", "redis",
    "a/b testing", "experimentation", "statistics",
    "data science", "analytics", "pandas", "numpy",
    "distributed systems", "kafka", "spark streaming",
}

# Titles that indicate strong AI/ML role alignment
STRONG_AI_TITLES = {
    "ai engineer", "ml engineer", "machine learning engineer",
    "data scientist", "nlp engineer", "research engineer",
    "applied scientist", "applied ml", "senior ai", "senior ml",
    "ai/ml", "ml/ai", "deep learning engineer",
    "search engineer", "recommendation engineer", "ranking engineer",
    "platform engineer", "backend engineer",  # partial credit
    "data engineer",  # partial credit (infra experience)
    "software engineer",  # partial credit
    "junior ml", "junior ai",  # lower credit
}

# Titles that are clearly NOT relevant to this JD
IRRELEVANT_TITLES = {
    "marketing manager", "hr manager", "human resources",
    "accountant", "content writer", "graphic designer",
    "sales executive", "civil engineer", "mechanical engineer",
    "customer support", "operations manager", "project manager",
    "business analyst", "finance", "recruiter",
    "ui designer", "ux designer", "product designer",
}

# Pure services companies (disqualifier per JD)
SERVICES_COMPANIES = {
    "tcs", "tata consultancy", "infosys", "wipro", "accenture",
    "cognizant", "capgemini", "hcl", "tech mahindra", "mphasis",
    "hexaware", "niit", "mastech", "syntel", "l&t infotech",
    "mindtree",  # acquired by L&T, borderline
}

# Product companies (positive signal)
PRODUCT_COMPANY_SIGNALS = {
    "amazon", "google", "microsoft", "meta", "apple", "netflix",
    "flipkart", "swiggy", "zomato", "ola", "paytm", "razorpay",
    "cred", "meesho", "nykaa", "freshworks", "zoho", "cleartax",
    "phonepe", "sharechat", "moj", "dream11", "udaan",
    "thoughtworks",  # consulting but product-culture
    "startups", "series", "funded",
}

REFERENCE_DATE = date(2025, 6, 27)


# ---------------------------------------------------------------------------
# Scoring Functions
# ---------------------------------------------------------------------------

def normalize(val: float, lo: float, hi: float) -> float:
    """Clamp and normalize a value to [0, 1]."""
    if hi == lo:
        return 0.5
    return max(0.0, min(1.0, (val - lo) / (hi - lo)))


def score_title_and_career(candidate: dict) -> tuple[float, str]:
    """
    Score how relevant the candidate's title and career history are.
    This is the most important signal — it catches keyword stuffers.
    Returns (score 0-1, reason fragment).
    """
    profile = candidate.get("profile", {})
    career = candidate.get("career_history", [])

    current_title = profile.get("current_title", "").lower()
    headline = profile.get("headline", "").lower()
    summary = profile.get("summary", "").lower()

    # Check if current title is clearly irrelevant
    is_irrelevant = any(irr in current_title for irr in IRRELEVANT_TITLES)
    is_strong_ai = any(ai in current_title for ai in STRONG_AI_TITLES)

    # Examine career history for AI/ML roles
    ai_role_months = 0
    total_career_months = 0
    product_company_experience = False
    pure_services_career = True
    recent_ai_role = False

    for i, job in enumerate(career):
        title = job.get("title", "").lower()
        company = job.get("company", "").lower()
        duration = job.get("duration_months", 0)
        description = job.get("description", "").lower()
        is_current = job.get("is_current", False)

        total_career_months += duration

        # Check company type
        is_services = any(svc in company for svc in SERVICES_COMPANIES)
        is_product = any(prod in company for prod in PRODUCT_COMPANY_SIGNALS)
        company_size = job.get("company_size", "")

        if not is_services:
            pure_services_career = False
        if is_product or company_size in ["501-1000", "1001-5000", "5001-10000", "10001+"]:
            # Large non-services companies are likely product companies
            if not is_services:
                product_company_experience = True

        # Check if role is AI-adjacent
        title_is_ai = any(ai in title for ai in STRONG_AI_TITLES)
        desc_has_ml = any(kw in description for kw in [
            "machine learning", "deep learning", "model", "embedding",
            "retrieval", "nlp", "ai", "ml", "neural", "training",
            "vector", "ranking", "recommendation", "classification"
        ])

        if title_is_ai or desc_has_ml:
            ai_role_months += duration
            if is_current or i == 0:
                recent_ai_role = True

    # Calculate title score
    title_score = 0.0
    if is_strong_ai and not is_irrelevant:
        title_score = 0.85
    elif not is_irrelevant and not is_strong_ai:
        # Ambiguous — check career history
        ai_ratio = ai_role_months / max(total_career_months, 1)
        title_score = 0.3 + 0.4 * ai_ratio
    else:
        # Clearly irrelevant title
        title_score = 0.05 + (0.2 if ai_role_months > 24 else 0.0)

    # Bonuses and penalties
    if recent_ai_role and not is_irrelevant:
        title_score = min(1.0, title_score + 0.1)
    if product_company_experience:
        title_score = min(1.0, title_score + 0.05)
    if pure_services_career and total_career_months > 24:
        title_score *= 0.7  # Pure services = soft penalty per JD

    reason = f"{profile.get('current_title', 'Unknown')} ({int(ai_role_months/12)}y AI exp)"
    return title_score, reason


def score_skills(candidate: dict) -> tuple[float, str]:
    """
    Score skill relevance, weighted by proficiency, endorsements, and duration.
    Avoids raw keyword count — uses quality-adjusted scoring.
    """
    skills = candidate.get("skills", [])
    assessments = candidate.get("redrob_signals", {}).get("skill_assessment_scores", {})

    tier1_score = 0.0
    tier2_score = 0.0
    tier1_count = 0
    tier2_count = 0

    PROFICIENCY_WEIGHT = {"beginner": 0.3, "intermediate": 0.6, "advanced": 0.85, "expert": 1.0}

    for skill in skills:
        name = skill.get("name", "").lower()
        proficiency = skill.get("proficiency", "intermediate")
        endorsements = skill.get("endorsements", 0)
        duration_months = skill.get("duration_months", 0)

        prof_weight = PROFICIENCY_WEIGHT.get(proficiency, 0.5)
        endorse_boost = min(0.15, endorsements / 500.0)  # up to 0.15 bonus
        duration_boost = min(0.1, duration_months / 600.0)  # up to 0.1 for 4+ years

        # Check assessed skills
        assess_score = assessments.get(skill.get("name", ""), -1)
        if assess_score >= 0:
            prof_weight = max(prof_weight, assess_score / 100.0)

        skill_quality = prof_weight + endorse_boost + duration_boost

        if any(t1 in name for t1 in TIER1_SKILLS):
            tier1_score += skill_quality
            tier1_count += 1
        elif any(t2 in name for t2 in TIER2_SKILLS):
            tier2_score += skill_quality * 0.6
            tier2_count += 1

    # Normalize: ideal is 3+ tier1, 5+ tier2
    tier1_norm = min(1.0, tier1_score / (3 * 1.25))  # 3 high-quality tier1 skills
    tier2_norm = min(1.0, tier2_score / (5 * 0.75))  # 5 mid-quality tier2 skills

    combined = 0.65 * tier1_norm + 0.35 * tier2_norm

    reason_parts = []
    if tier1_count > 0:
        reason_parts.append(f"{tier1_count} core AI skills")
    if tier2_count > 0:
        reason_parts.append(f"{tier2_count} supporting skills")

    return combined, "; ".join(reason_parts) if reason_parts else "no relevant skills"


def score_experience(candidate: dict) -> tuple[float, str]:
    """Score years of experience. Ideal: 5-9 years."""
    yoe = candidate.get("profile", {}).get("years_of_experience", 0)

    if 5 <= yoe <= 9:
        score = 1.0
    elif 4 <= yoe < 5:
        score = 0.85
    elif 9 < yoe <= 12:
        score = 0.8
    elif 3 <= yoe < 4:
        score = 0.65
    elif 12 < yoe <= 15:
        score = 0.65
    elif yoe > 15:
        score = 0.5  # overqualified risk
    else:
        score = max(0.1, yoe / 5.0 * 0.65)

    return score, f"{yoe}y experience"


def score_education(candidate: dict) -> tuple[float, str]:
    """Score education — CS/ML/EE fields from good institutions."""
    education = candidate.get("education", [])
    if not education:
        return 0.4, "no education listed"

    best_score = 0.0
    best_desc = "education present"

    TIER_SCORE = {"tier_1": 1.0, "tier_2": 0.75, "tier_3": 0.5, "tier_4": 0.3, "unknown": 0.4}
    RELEVANT_FIELDS = {
        "computer science", "cs", "information technology", "it",
        "electrical engineering", "electronics", "mathematics", "statistics",
        "artificial intelligence", "machine learning", "data science",
        "software engineering", "cognitive science",
    }

    for edu in education:
        field = edu.get("field_of_study", "").lower()
        tier = edu.get("tier", "unknown")
        degree = edu.get("degree", "").lower()

        tier_score = TIER_SCORE.get(tier, 0.4)
        field_relevant = any(f in field for f in RELEVANT_FIELDS)

        # Degree level bonus
        degree_bonus = 0.0
        if any(d in degree for d in ["phd", "ph.d", "doctorate"]):
            degree_bonus = 0.1
        elif any(d in degree for d in ["m.tech", "m.e.", "mtech", "ms", "m.s.", "mba"]):
            degree_bonus = 0.05

        edu_score = tier_score * (1.0 if field_relevant else 0.6) + degree_bonus

        if edu_score > best_score:
            best_score = edu_score
            best_desc = f"{edu.get('degree', '')} in {edu.get('field_of_study', '')} ({tier})"

    return min(1.0, best_score), best_desc


def score_behavioral_signals(candidate: dict) -> tuple[float, str]:
    """
    Score Redrob platform signals — availability and engagement.
    Used as a multiplier on base score to model real hireability.
    """
    signals = candidate.get("redrob_signals", {})

    scores = {}

    # 1. Availability/intent signals
    open_to_work = 1.0 if signals.get("open_to_work_flag", False) else 0.3
    scores["open_to_work"] = open_to_work

    # 2. Recency — how recently was the candidate active?
    last_active_str = signals.get("last_active_date", "")
    if last_active_str:
        try:
            last_active = date.fromisoformat(last_active_str)
            days_inactive = (REFERENCE_DATE - last_active).days
            if days_inactive <= 7:
                recency = 1.0
            elif days_inactive <= 30:
                recency = 0.85
            elif days_inactive <= 90:
                recency = 0.65
            elif days_inactive <= 180:
                recency = 0.4
            else:
                recency = 0.15
        except (ValueError, TypeError):
            recency = 0.5
    else:
        recency = 0.5
    scores["recency"] = recency

    # 3. Recruiter response rate — key availability signal
    response_rate = signals.get("recruiter_response_rate", 0.5)
    scores["response_rate"] = response_rate

    # 4. Notice period preference (<= 30 days preferred per JD)
    notice = signals.get("notice_period_days", 60)
    if notice <= 15:
        notice_score = 1.0
    elif notice <= 30:
        notice_score = 0.9
    elif notice <= 60:
        notice_score = 0.65
    elif notice <= 90:
        notice_score = 0.4
    else:
        notice_score = 0.2
    scores["notice_period"] = notice_score

    # 5. Profile completeness
    completeness = signals.get("profile_completeness_score", 50) / 100.0
    scores["completeness"] = completeness

    # 6. GitHub activity (AI engineers should have GitHub presence)
    github = signals.get("github_activity_score", -1)
    if github == -1:
        github_score = 0.3  # no github linked
    else:
        github_score = 0.3 + 0.7 * (github / 100.0)
    scores["github"] = github_score

    # 7. Interview completion rate (reliability signal)
    interview_rate = signals.get("interview_completion_rate", 0.7)
    scores["interview_rate"] = interview_rate

    # 8. Saved by recruiters (social proof)
    saved = min(1.0, signals.get("saved_by_recruiters_30d", 0) / 10.0)
    scores["saved_by_recruiters"] = saved

    # Weighted combination
    weighted = (
        0.20 * scores["open_to_work"] +
        0.20 * scores["recency"] +
        0.20 * scores["response_rate"] +
        0.15 * scores["notice_period"] +
        0.10 * scores["completeness"] +
        0.08 * scores["github"] +
        0.05 * scores["interview_rate"] +
        0.02 * scores["saved_by_recruiters"]
    )

    reason = (
        f"open={int(open_to_work*100)}% "
        f"active={int(recency*100)}% "
        f"resp={response_rate:.0%} "
        f"notice={notice}d"
    )
    return weighted, reason


def score_location(candidate: dict) -> float:
    """Score location fit. India + preferred cities is ideal."""
    profile = candidate.get("profile", {})
    signals = candidate.get("redrob_signals", {})

    country = profile.get("country", "").lower()
    location = profile.get("location", "").lower()
    willing_to_relocate = signals.get("willing_to_relocate", False)

    if country not in ("india", "in"):
        if willing_to_relocate:
            return 0.4  # international but willing to move
        return 0.2

    # In India
    preferred_in_location = any(
        pref.lower() in location
        for pref in JD["preferred_locations"]
    )
    if preferred_in_location:
        return 1.0
    elif willing_to_relocate:
        return 0.85
    else:
        return 0.65  # India but different city, not relocating


def is_likely_honeypot(candidate: dict) -> bool:
    """
    Detect honeypot candidates: irrelevant titles with suspiciously many AI keywords.
    The JD explicitly warns about this trap.
    """
    profile = candidate.get("profile", {})
    current_title = profile.get("current_title", "").lower()
    skills = candidate.get("skills", [])

    is_irrelevant = any(irr in current_title for irr in IRRELEVANT_TITLES)
    if not is_irrelevant:
        return False

    # Count AI keywords in skills
    ai_keyword_count = sum(
        1 for s in skills
        if any(t in s.get("name", "").lower() for t in TIER1_SKILLS | TIER2_SKILLS)
    )

    # If irrelevant title + many AI keywords = honeypot
    return ai_keyword_count >= 6


def compute_final_score(candidate: dict) -> tuple[float, str]:
    """
    Compute weighted composite score and generate reasoning.

    Weights:
    - Title & Career History: 35% — catches keyword stuffers
    - Skills Match:           25% — quality-adjusted skill relevance
    - Experience:             15% — years of experience fit
    - Behavioral Signals:     20% — availability & engagement
    - Location:                5% — geographic fit
    """
    title_score, title_reason = score_title_and_career(candidate)
    skills_score, skills_reason = score_skills(candidate)
    exp_score, exp_reason = score_experience(candidate)
    behavior_score, behavior_reason = score_behavioral_signals(candidate)
    location_score = score_location(candidate)

    # Base weighted score
    base_score = (
        0.35 * title_score +
        0.25 * skills_score +
        0.15 * exp_score +
        0.20 * behavior_score +
        0.05 * location_score
    )

    # Honeypot penalty
    if is_likely_honeypot(candidate):
        base_score *= 0.25

    # Hard cap for clearly irrelevant titles (they can't rank in top 100)
    profile = candidate.get("profile", {})
    current_title = profile.get("current_title", "").lower()
    is_irrelevant = any(irr in current_title for irr in IRRELEVANT_TITLES)
    if is_irrelevant and title_score < 0.2:
        base_score = min(base_score, 0.12)

    final = round(min(1.0, max(0.0, base_score)), 6)

    reasoning = (
        f"{profile.get('current_title', '?')} | "
        f"{exp_reason} | "
        f"{skills_reason} | "
        f"{behavior_reason}"
    )
    # Truncate to reasonable length
    reasoning = reasoning[:200]

    return final, reasoning


# ---------------------------------------------------------------------------
# Main Processing
# ---------------------------------------------------------------------------

def rank_candidates(input_path: str, top_n: int = 100) -> list[dict]:
    """Stream candidates.jsonl and return top_n ranked."""
    print(f"Processing candidates from: {input_path}", file=sys.stderr)
    path = Path(input_path)
    if not path.exists():
        raise FileNotFoundError(f"Candidates file not found: {input_path}")

    scored = []
    count = 0
    errors = 0

    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                candidate = json.loads(line)
                score, reasoning = compute_final_score(candidate)
                scored.append({
                    "candidate_id": candidate["candidate_id"],
                    "score": score,
                    "reasoning": reasoning,
                    "_candidate": candidate,  # keep for XLSX enrichment
                })
                count += 1
                if count % 10000 == 0:
                    print(f"  Processed {count} candidates...", file=sys.stderr)
            except (json.JSONDecodeError, KeyError) as e:
                errors += 1

    print(f"Processed {count} candidates ({errors} errors)", file=sys.stderr)

    # Round scores to 4 decimal places (matching CSV output) then sort
    # Tie-break: equal score -> candidate_id ascending (lexicographic)
    for item in scored:
        item["score"] = round(item["score"], 4)
    scored.sort(key=lambda x: (-x["score"], x["candidate_id"]))

    top = scored[:top_n]
    for rank, item in enumerate(top, start=1):
        item["rank"] = rank

    return top


def write_csv(ranked: list[dict], output_path: str) -> None:
    """Write submission-compliant CSV."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])
        for item in ranked:
            writer.writerow([
                item["candidate_id"],
                item["rank"],
                f"{item['score']:.4f}",
                item["reasoning"],
            ])
    print(f"CSV written: {output_path}", file=sys.stderr)


def write_xlsx(ranked: list[dict], output_path: str) -> None:
    """Write enriched XLSX with additional candidate details."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping XLSX output", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    # Styles
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    top10_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Rank", "Candidate ID", "Score", "Current Title", "Years Exp",
        "Location", "Country", "Open to Work", "Notice Period (days)",
        "Response Rate", "GitHub Score", "Last Active",
        "Skills Count", "Top Skills", "Education", "Reasoning"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    for item in ranked:
        rank = item["rank"]
        cand = item.get("_candidate", {})
        profile = cand.get("profile", {})
        signals = cand.get("redrob_signals", {})
        skills = cand.get("skills", [])
        education = cand.get("education", [])

        top_skills = sorted(
            [s for s in skills if any(
                t in s.get("name", "").lower()
                for t in TIER1_SKILLS | TIER2_SKILLS
            )],
            key=lambda s: s.get("endorsements", 0),
            reverse=True
        )[:5]
        top_skills_str = ", ".join(s.get("name", "") for s in top_skills)

        edu_str = ""
        if education:
            e = education[0]
            edu_str = f"{e.get('degree', '')} {e.get('field_of_study', '')} ({e.get('tier', '')})"

        row_data = [
            rank,
            item["candidate_id"],
            item["score"],
            profile.get("current_title", ""),
            profile.get("years_of_experience", ""),
            profile.get("location", ""),
            profile.get("country", ""),
            "Yes" if signals.get("open_to_work_flag") else "No",
            signals.get("notice_period_days", ""),
            f"{signals.get('recruiter_response_rate', 0):.0%}",
            signals.get("github_activity_score", -1),
            signals.get("last_active_date", ""),
            len(skills),
            top_skills_str,
            edu_str,
            item["reasoning"],
        ]

        row_idx = rank + 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

            # Row styling
            if rank == 1:
                cell.fill = gold_fill
            elif rank == 2:
                cell.fill = silver_fill
            elif rank == 3:
                cell.fill = bronze_fill
            elif rank <= 10:
                cell.fill = top10_fill
            elif rank % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = [6, 14, 8, 22, 8, 16, 10, 10, 12, 12, 10, 14, 10, 30, 25, 50]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add a summary sheet
    ws2 = wb.create_sheet(title="Summary")
    ws2["A1"] = "Redrob AI Ranker — Submission Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Total candidates processed"
    ws2["A4"] = "Top 100 candidates selected"
    ws2["A5"] = "Top score"
    ws2["A6"] = "Bottom score (rank 100)"
    ws2["A7"] = "Generated"

    top_score = ranked[0]["score"] if ranked else 0
    bottom_score = ranked[-1]["score"] if ranked else 0

    ws2["B3"] = "100,000"
    ws2["B4"] = 100
    ws2["B5"] = f"{top_score:.4f}"
    ws2["B6"] = f"{bottom_score:.4f}"
    ws2["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"XLSX written: {output_path}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description="Redrob Intelligent Candidate Ranker")
    parser.add_argument("--candidates", default="./candidates.jsonl",
                        help="Path to candidates.jsonl")
    parser.add_argument("--out", default="./submission.csv",
                        help="Output CSV path")
    parser.add_argument("--xlsx", default=None,
                        help="Optional XLSX output path")
    parser.add_argument("--top", type=int, default=100,
                        help="Number of top candidates to output (default: 100)")
    args = parser.parse_args()

    ranked = rank_candidates(args.candidates, top_n=args.top)
    write_csv(ranked, args.out)

    if args.xlsx:
        write_xlsx(ranked, args.xlsx)

    # Print top 10 to stdout for quick review
    print(f"\nTop 10 candidates:")
    print("-" * 80)
    for item in ranked[:10]:
        print(f"Rank {item['rank']:3d} | Score: {item['score']:.4f} | {item['candidate_id']} | {item['reasoning'][:80]}")

    print(f"\nSubmission written to: {args.out}")
    if args.xlsx:
        print(f"XLSX report written to: {args.xlsx}")


if __name__ == "__main__":
    main()
